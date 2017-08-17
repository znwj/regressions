#!/usr/bin/python  
#-*- coding:utf-8 -*-  
############################  
#File Name: getlabel.py
#Author: Wenjie Zhang 
#Mail: zwjhit@gmail.com  
#Created Time: 2017-08-09 10:15:38
############################  

from openpyxl.reader.excel import load_workbook
import numpy as np
import pickle

def score1(label,data):
    precision=[]
    for i in range(4):
        precision.append(sum(data[:,i]*label[:,i])/sum(data[:,i]))
    artifact_sensitivity=sum(data[:,3]*label[:,3])/sum(label[:,3])
    artifact_specificity=sum(sum(data[:,0:3]*label[:,0:3]))/36
    heartproblem_detection_sensitivity=sum(sum(data[:,1:3]*label[:,1:3]))/22
    heartproblem_detection_precision=sum(sum(data[:,1:3]*label[:,1:3]))/sum(sum(data[:,1:3]))
    y_index=sum(data[:,3]*label[:,3])/16-(1-artifact_specificity)
    f_score=1.81*heartproblem_detection_sensitivity*heartproblem_detection_precision/(
    0.81*heartproblem_detection_precision+heartproblem_detection_sensitivity)
    total_score=sum(precision)
    normal_score=(precision[0]*14+precision[1]*14+precision[2]*8+precision[3]*16)/52

    score_result=[]
    score_result.append({'Precision of Normal':precision[0]})
    score_result.append({'Precision of Murmur':precision[1]})
    score_result.append({'Precision of Extra Sound':precision[2]})
    score_result.append({'Precision of Artifact':precision[3]})
    score_result.append({'Artifact Sensitivity':artifact_sensitivity})
    score_result.append({'Artifact Specificity':artifact_specificity})
    score_result.append({'Heartproblem Detection Sensitivity':heartproblem_detection_sensitivity})
    score_result.append({'Heartproblem Detection Precision':heartproblem_detection_precision})
    score_result.append({'Youden Index of Artifact':y_index})
    score_result.append({'F-Score of Heartproblem Detection':f_score})
    score_result.append({'Total Precision':total_score})
    score_result.append({'Normalized Precision':normal_score})

    return score_result

def score2(label,data):
    precision=[]
    for i in range(3):
        precision.append(sum(data[:,i]*label[:,i])/sum(data[:,i]))
    sensitivity_of_heartproblem=sum(sum(data[:,1:3]*label[:,1:3]))/59
    specificity_of_heartproblem=sum(data[:,0]*label[:,0])/136
    y_index=sensitivity_of_heartproblem+specificity_of_heartproblem-1
    dp=(np.sqrt(3)/np.pi)*(np.log10(sensitivity_of_heartproblem/(1-sensitivity_of_heartproblem))
        +np.log10(specificity_of_heartproblem/(1-specificity_of_heartproblem)))
    total_score=sum(precision)
    normal_score=(precision[0]*136+precision[1]*39+precision[2]*20)/195

    score_result=[]
    score_result.append({'Precision of Normal':precision[0]})
    score_result.append({'Precision of Murmur':precision[1]})
    score_result.append({'Precision of Extrastole':precision[2]})
    score_result.append({'Sensitivity of Heartproblem':sensitivity_of_heartproblem})
    score_result.append({'Specificity of Heartproblem':specificity_of_heartproblem})
    score_result.append({'Youden Index of Artifact':y_index})
    score_result.append({'Discriminant Power':dp})
    score_result.append({'Total Precision':total_score})
    score_result.append({'Normalized Precision':normal_score})
    
    return score_result

if __name__=="__main__":
    wb=load_workbook("1.xlsx")
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])

    label1=np.zeros((52,4))
    title=['B','C','D','E']
    for i in range(4):
        for j in range(52):
            label1[j,i]=ws.cell(title[i]+str(55+j)).value
    #print (label1)

    ws2 = wb.get_sheet_by_name(sheetnames[1])
    label2=np.zeros((195,3))
    for i in range(3):
        for j in range(195):
            label2[j,i]=ws2.cell(title[i]+str(198+j)).value
    #print(label2)

    data1=np.zeros((52,4))
    for i in range(4):
        for j in range(52):
            data1[j,i]=ws.cell(title[i]+str(2+j)).value

    data2=np.zeros((195,3))
    for i in range(3):
        for j in range(195):
            data2[j,i]=ws2.cell(title[i]+str(2+j)).value
    
    score_result1=score1(label1,data1)
    for i in score_result1:
        print(i)
    score_result2=score2(label2,data2)
    for i in score_result2:
        print(i)
    
    f=open("label.dat","wb")
    pickle.dump([label1,label2],f)
    f.close()
    f2=open("label.dat","rb")
    c=pickle.load(f2)
